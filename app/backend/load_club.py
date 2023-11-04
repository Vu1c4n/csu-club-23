import pymongo
from openpyxl import load_workbook
from cfg import *

myclient = pymongo.MongoClient(**database_cfg)
mydb = myclient['club23']
club = mydb['all_club']

wb = load_workbook("./static/社团名单.xlsx")
ws = wb.active
start = 1
lunx = 0
x = {"_id": "allclub"}
while True:
    if (ws[f'B{start}'].value != None):
        if (ws[f'B{start}'].value != "名称"):
            x[str(lunx)] = ws[f'B{start}'].value
            start += 1
            lunx += 1
        else:
            print("跳过")
            start += 1
            continue
    else:
        break
club.insert_one(x)
print(x)
