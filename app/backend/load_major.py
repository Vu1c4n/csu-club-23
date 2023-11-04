from openpyxl import load_workbook
import json
from db import apply
wb = load_workbook("./static/2022级学院大类名单.xlsx")
ws = wb.active
start = 2
all = []
children = []
x = ""
while (ws[f'B{start}'].value != None):
    if ws[f'B{start}'].value != x:
        if x != "":
            all.append({"text": x, "value": str(start + 10000), "children": children})
            children = []
            x = ws[f'B{start}'].value
            children.append({"text": ws[f'A{start}'].value, "value": start})
        else:
            x = ws[f'B{start}'].value
            children.append({"text": ws[f'A{start}'].value, "value": start})
        start += 1
    else:
        children.append({"text": ws[f'A{start}'].value, "value": start})
        start += 1

all.append({"text": "生命科学学院", "value": "13543", "children": [{"text": "生物科学拔尖人才培养", "value": "62378"}]})

apply.insert_one({"_id":"options-22","data":all}) # 将选项加入列表