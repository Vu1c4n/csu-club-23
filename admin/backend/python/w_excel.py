# w_excel.py 封装excel操作

from openpyxl import load_workbook,Workbook
from cfg import DATA_PATH

# private member
__work_book = load_workbook(filename= DATA_PATH + 'template/tmpl.xlsx')
__ws = __work_book.active
__NAME_CELL = (2,2)
__STU_START_ROW = 5

def __cell(x:int,y:int):
    return __ws.cell(x,y)

def __write_club_name(name:str):
    name_cell = __cell(*__NAME_CELL)
    name_cell.value = name

def __write_stu_one(row, info):
    if (info == None) or ('name' not in info.keys()):
        return 1
    __cell(row,2).value = info['name']
    __cell(row,3).value = info['sex']
    __cell(row,4).value = info['nation']
    __cell(row,5).value = info['birth']
    __cell(row,6).value = info['political']
    __cell(row,7).value = info['stu_type']
    __cell(row,8).value = info['stu_id']
    __cell(row,9).value = info['academy']
    __cell(row,10).value = info['specialty']
    __cell(row,11).value = info['grade']
    __cell(row,12).value = info['stu_class']
    __cell(row,13).value = None
    __cell(row,14).value = info['phone']
    __cell(row,15).value = info['qq']
    __cell(row,16).value = info['from_location']

def __write_stu_all(s_list:list,c_name:str):
    row = __STU_START_ROW
    for stu in s_list:
        __write_stu_one(row,stu)
        __cell(row,13).value = c_name # 填写“所在社团”字段，不规范的写法，缘于主服务缺陷
        row += 1


# APIs
def write(c_name,s_list):
    __write_club_name(c_name)
    __write_stu_all(s_list,c_name)
    __work_book.save(DATA_PATH + 'output/'+ c_name + '.xlsx')


# unit_test
if __name__ == '__main__':
    dt = {'name': 'xxx', 'sex': '男', 'nation': '汉族', 'birth': '2004年4月', 'political': '共青团员', 'stu_id': '8207221323', 'stu_type': '本科', 'specialty': '人工智能', 'grade': '2022级', 'academy': '自动化学院', 'stu_class': '智能2202', 'from_location': '中国内地（大陆)', 'phone': '19307491691', 'qq': '1366025392'}
    write('法鹰社',[dt,dt])
